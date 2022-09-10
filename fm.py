import openpyxl
from openpyxl.utils.exceptions import SheetTitleException
import os
import re
import DTC


class TestCaseFM:
    # def __init__(self, Extra):
    #     self.Path_csv = ''
    #     self.PathSystemDegradation = ''
    #     self.PathTemplate = ''
    #     self.PathBook1 = Extra

    # FCO_Position = 'Z3'

    # First_Fault_Position = 'C5'

    # Path_csv = r'D:\SandBox\Clara\BORGWARD\S701\01_EventId_ITC_index_table_DEM_DDL.csv'
    # PathSystemDegradation = r'D:\SandBox\Clara\BORGWARD\S701\Borgward_S701_System_Degradation_Strategy_Internal_Table_v2.0_17_Dec_2019.xlsx'
    # PathTemplate = r'C:\Users\udn1hc\PycharmProjects\untitled\data\Template.py'
    # PathBook1 = r'D:\SandBox\Clara\BORGWARD\S701\Foton_S701.xlsx'
    # PathDBC = 'D:\SandBox\Clara\BORGWARD\S701\SYSB_PLUS.FOTON_BORGWARD.S701-R5.0_I3_FOTON_BORGWARD_S701\RadarFC\measurement\database\dbc\BORGWARD_S701_PUB.dbc'
    PathOfActivateFile = 'output/_prj_para'
    # sheet_name = 'System Degradation Strategy'
    # PathventID = r'D:\Tool\CLARA_Tool\Develop\CN202\Dem_EventIds_Project.h'
    # Video = 'AWV_VRUWARNINGFUNCTION_OFF'
    Confirm_PF_28 = False

    List_Fault = []
    List_Fault_Alive = []
    List_Fault_Timeout = []
    List_Fault_CheckSum = []
    List_Fault_DLC = []
    List_Fault_Signal = []
    List_Fault_PF_MM = []
    DTCDict = {}
    FromTem = ''
    DataPF = ['CANSM', 'ALIGNMENT', 'PRODUCTION', 'SENSOR_BLIND', 'SYSTEM_BLIND', 'SYSTEM_DISTURBED', 'VBAT', 'CAN',
              'DEFECT',
              'RADAR_UNAVAILABLE', 'MISALIGNMENT', 'VIDEO_SIGNAL', 'FREE_SIGHT', 'XML']
    count = 0
    count2 = 0
    Max_Row_System_Degradation = 0
    FCO_Column = 0
    FCS_Column = 0
    sheet = None

    def GetColumn_Number(self):
        listCol = {'A': 1,
                   'B': 2,
                   'C': 3,
                   'D': 4,
                   'E': 5,
                   'F': 6,
                   'G': 7,
                   'H': 8,
                   'I': 9,
                   'J': 10,
                   'K': 11,
                   'L': 12,
                   'M': 13,
                   'N': 14,
                   'O': 15,
                   'P': 16,
                   'Q': 17,
                   'R': 18,
                   'S': 19,
                   'T': 20,
                   'U': 21,
                   'V': 22,
                   'W': 23,
                   'X': 24,
                   'Y': 25,
                   'Z': 26,
                   }
        return listCol[self]

    def GetHex(self):
        listint = {
            '1': 'A',
            '2': 'B',
            '3': 'C',
            '4': 'D',
            '5': 'E',
            '6': 'F',
            '7': 'G',
            '8': 'H',
            '9': 'I',
            '10': 'J',
            '11': 'K',
            '12': 'L',
            '13': 'M',
            '14': 'N',
            '15': 'O',
            '16': 'P',
            '17': 'Q',
            '18': 'R',
            '19': 'S',
            '20': 'T',
            '21': 'U',
            '22': 'V',
            '23': 'W',
            '24': 'X',
            '25': 'Y',
            '26': 'Z',
        }
        return listint[self]

    def Openxlsx(self, First_Fault_Position, sheet_name, FCO_Position):
        book = openpyxl.load_workbook(self)
        # sheet = book['CDM']
        # TestCaseFM.sheet = book.active

        try:
            TestCaseFM.sheet = book[sheet_name]
        except SheetTitleException:
            TestCaseFM.sheet = book.active
        # sheet.max_column
        t = TestCaseFM.sheet.max_row
        i = int(First_Fault_Position[1])

        for j in range(i, t):
            if TestCaseFM.sheet.cell(row=j, column=TestCaseFM.GetColumn_Number(str(First_Fault_Position[0]).upper())).value is not None:
                TestCaseFM.count2 += 1
            else:
                break
        TestCaseFM.Max_Row_System_Degradation = TestCaseFM.count2 + i

        tt_FCO_Column = 0
        t_FCO_Len = 0
        t_FCS_Position = ''
        t_FCS_Len = 0
        check_position = ''
        check_len = ''
        track_FCS = 0

        if len(FCO_Position) > 2:
            tt_FCO_Column = TestCaseFM.GetColumn_Number(str(FCO_Position[:1]).upper()) * 26
            tt_FCO_Column = tt_FCO_Column + TestCaseFM.GetColumn_Number(str(FCO_Position[1:2]).upper())
            print('AAAAAA', tt_FCO_Column)
        else:
            tt_FCO_Column = TestCaseFM.GetColumn_Number(str(FCO_Position[0]).upper())

        print(TestCaseFM.sheet.cell(row=int(FCO_Position[-1]), column=tt_FCO_Column).value)
        check_len = TestCaseFM.sheet.cell(int(FCO_Position[-1]), tt_FCO_Column).value
        while check_len == 'FCO' or check_len == 'FCS' or check_len is None:
            if check_len == 'FCO':
                check_position = 'FCO'
            elif check_len == 'FCS':
                check_position = 'FCS'
                track_FCS = tt_FCO_Column

            if (check_len == 'FCO' or check_len is None) and check_position != 'FCS':
                t_FCO_Len += 1
                tt_FCO_Column += 1
                check_len = TestCaseFM.sheet.cell(int(FCO_Position[-1]), tt_FCO_Column).value
            elif (check_len == 'FCS' or check_len is None) and check_position != 'FCO':
                t_FCS_Len += 1
                tt_FCO_Column += 1
                check_len = TestCaseFM.sheet.cell(int(FCO_Position[-1]), tt_FCO_Column).value
        print(t_FCO_Len, t_FCS_Len)

        for i in range(5):
            if track_FCS <= 26 * i:
                if i > 1:
                    t_track_FCS = track_FCS - 26 * (i - 1)
                    t_FCS_Position = TestCaseFM.GetHex(str(i - 1))
                    t_FCS_Position += TestCaseFM.GetHex(str(t_track_FCS))
                    tt = FCO_Position[-1]
                    t_FCS_Position += str(int(FCO_Position[-1]) + 2)
                else:
                    t_FCS_Position = TestCaseFM.GetHex(str(track_FCS))
                    t_FCS_Position += str(int(FCO_Position[-1]) + 2)
                break

        return t_FCO_Len, t_FCS_Len, t_FCS_Position

    def Countcsv(self):
        gendtc = open(self, "r")
        for line in gendtc:
            if line != '':
                TestCaseFM.count += 1
            else:
                break
        gendtc.close()

    def GenDTCDict(self):
        gendtc = open(self, "r")
        for i in range(TestCaseFM.count):
            DTCDict_specil = {}
            strr = gendtc.readline().split(',')[0]
            if i == 0:
                continue
            else:
                str1 = strr.split(";")
                if '0x' in str1[1]:
                    TestCaseFM.Confirm_PF_28 = True
                    if str1[0] != '':
                        str1[0] = 'DTC_' + str1[0].strip()
                        str1[1] = str1[1].strip()
                        str1[3] = str1[3].strip()
                        if len(str1[0]) > 5 and str1[0][:4] == "DTC_":
                            TestCaseFM.List_Fault.append(str1[3])
                            # messagebox.showinfo("Title", TestCaseFM.List_Fault)
                            DTCDict_specil[str1[3]] = {}
                            DTCDict_specil[str1[3]]['Fault_ID'] = 'AAAA'
                            DTCDict_specil[str1[3]]['DTC'] = str1[0]
                            DTCDict_specil[str1[3]]['FUNCFCO'] = ''
                            DTCDict_specil[str1[3]]['FUNCFCS'] = ''
                            DTCDict_specil[str1[3]]['MAXTIME'] = ''
                            DTCDict_specil[str1[3]]['MINTIME'] = ''
                            DTCDict_specil[str1[3]]['MESSAGE'] = ''
                            DTCDict_specil[str1[3]]['SIGNAL'] = ''
                            TestCaseFM.DTCDict.update(DTCDict_specil)
                else:
                    if len(str1[2]) > 5 and str1[2][:4] == "DTC_":
                        TestCaseFM.List_Fault.append(str1[0])
                        DTCDict_specil[str1[0].strip()] = {}
                        DTCDict_specil[str1[0]]['Fault_ID'] = str1[1]
                        DTCDict_specil[str1[0]]['DTC'] = str1[2]
                        DTCDict_specil[str1[0]]['FUNCFCO'] = ''
                        DTCDict_specil[str1[0]]['FUNCFCS'] = ''
                        DTCDict_specil[str1[0]]['MAXTIME'] = ''
                        DTCDict_specil[str1[0]]['MINTIME'] = ''
                        DTCDict_specil[str1[0]]['MESSAGE'] = ''
                        DTCDict_specil[str1[0]]['SIGNAL'] = ''
                        TestCaseFM.DTCDict.update(DTCDict_specil)
        print(len(TestCaseFM.DTCDict), TestCaseFM.List_Fault)
        gendtc.close()

    def ReadFormTem(self):
        Templatebegin = open(self, 'r')
        TestCaseFM.FromTem = Templatebegin.read()
        Templatebegin.close()

    def SetBit(self, value, matchbit):
        listbit = {0: 1,
                   1: 2,
                   2: 4,
                   3: 8,
                   4: 16,
                   5: 32,
                   6: 64,
                   7: 128,
                   8: 256,
                   9: 512
                   }
        value |= listbit[matchbit]
        return value

    def CalPosition(self, FCS_Position):
        t_FCO_Column = 0
        t_FCS_Column = 0
        if len(self) > 2:
            t_FCO_Column = TestCaseFM.GetColumn_Number(str(self[:1]).upper()) * 26
            t_FCO_Column = t_FCO_Column + TestCaseFM.GetColumn_Number(str(self[1:2]).upper())
            TestCaseFM.FCO_Column = t_FCO_Column
            # print(t_FCO_Column)
        else:
            TestCaseFM.FCO_Column = TestCaseFM.GetColumn_Number(str(self[0]).upper())

        if len(FCS_Position) > 2:
            t_FCS_Column = TestCaseFM.GetColumn_Number(str(FCS_Position[0:1]).upper()) * 26
            t_FCS_Column = t_FCS_Column + TestCaseFM.GetColumn_Number(str(FCS_Position[1:2]).upper())
            TestCaseFM.FCS_Column = t_FCS_Column
            # print(t_FCS_Column)
        else:
            TestCaseFM.FCS_Column = TestCaseFM.GetColumn_Number(str(FCS_Position[0]).upper())

    def FCO(self, FCO_Position, FCO_Len, First_Fault_Position):
        for fco in range(FCO_Len):
            t1 = int(FCO_Position[len(FCO_Position) - 1]) + 2
            t2 = self.FCO_Column + fco
            find_fco = TestCaseFM.sheet.cell(row=int(FCO_Position[len(FCO_Position) - 1]) + 2,
                                             column=self.FCO_Column + fco).value
            if find_fco in ['FCO_NOREACTION', 'AWV_NOREACTION']:
                row_fco = int(FCO_Position[len(FCO_Position) - 1]) + 2
                column_fco = self.FCO_Column + fco
                # print('Column FCO No_Reaction = {0}'.format(column_fco))  # 3 16
        # MAX_ROW = TestCaseFM.sheet.max_row + 1
        Position_Fault = int(First_Fault_Position[1])
        FAULT_FCO_REDUNDANCE = []
        for i in range(Position_Fault, self.Max_Row_System_Degradation):
            reaction_fco = []
            for j in range(FCO_Len):
                reaction_fco.append(TestCaseFM.sheet.cell(row=i, column=self.FCO_Column + j).value)
            # print(len(reaction_fco))
            Flag = ''
            for check_IR_RE in reaction_fco:
                if check_IR_RE == 'IR' or check_IR_RE == 'ir':
                    Flag = 'IR'
                elif Flag != 'IR' and (check_IR_RE == 'RE' or check_IR_RE == 're'):
                    Flag = 'RE'
                elif Flag != 'IR' and Flag != 'RE':
                    Flag = ''

            Fault_Name_FCO = TestCaseFM.sheet[First_Fault_Position[0] + '{0}'.format(i)].value.strip()
            try:
                if Flag == 'RE' and (reaction_fco[column_fco - self.FCO_Column] == 'RE' or reaction_fco[
                    column_fco - self.FCO_Column] == 're'):
                    TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_NoChange'
                elif Flag == 'IR':
                    TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_IRR_OFF'
                elif Flag == 'RE' and (reaction_fco[column_fco - self.FCO_Column] != 'RE' or reaction_fco[
                    column_fco - self.FCO_Column] != 're'):
                    TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_REV_OFF'
                else:
                    TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_NoChange'
            except:
                temp_fault = Fault_Name_FCO
                t_Red = re.search(r'(.*)FAULT_(.*)', temp_fault, re.M | re.I)
                if t_Red == None:
                    t_Red = re.search(r'(.*?)_(.*)', temp_fault, re.M | re.I)
                # print(t_Red.group(2))
                try:
                    for li_st in self.List_Fault:
                        if t_Red.group(2) in li_st[:]:
                            Fault_Name_FCO = li_st
                            break
                except:
                    print('Not find redundance {0}'.format(Fault_Name_FCO))
                # print(Fault_Name_FCO)
                try:
                    if Flag == 'RE' and (reaction_fco[column_fco - self.FCO_Column] == 'RE' or reaction_fco[
                        column_fco - self.FCO_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_NoChange'
                    elif Flag == 'IR':
                        TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_IRR_OFF'
                    elif Flag == 'RE' and (reaction_fco[column_fco - self.FCO_Column] != 'RE' or reaction_fco[
                        column_fco - self.FCO_Column] != 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_REV_OFF'
                    else:
                        TestCaseFM.DTCDict[str(Fault_Name_FCO)]['FUNCFCO'] = 'ACC_NoChange'
                except:
                    print('Can not add {0} to TestCase'.format(Fault_Name_FCO))

                FAULT_FCO_REDUNDANCE.append(temp_fault)
        Redundance = open(TestCaseFM.PathOfActivateFile + '\\Redundance.py', 'w')
        Redundance.write('"=============Fault in System degradation, but not in file.csv==========="' + '\n' * 2)
        for Re in FAULT_FCO_REDUNDANCE:
            Redundance.write(str(Re) + '\n')
        Redundance.close()
        # print(len(FAULT_FCO_REDUNDANCE), FAULT_FCO_REDUNDANCE)

    def FCS(self, FCS_Position, FCS_Len, First_Fault_Position, Video):
        for fcs in range(FCS_Len):
            find_fcs = TestCaseFM.sheet.cell(row=int(FCS_Position[len(FCS_Position) - 1]),
                                             column=self.FCS_Column + fcs).value
            if find_fcs in ['FCS_NOREACTION', 'AWV_NOREACTION']:
                row_fcs = int(FCS_Position[len(FCS_Position) - 1])
                column_fcs = self.FCS_Column + fcs
                print('Column FCS No_Reaction = {0}'.format(column_fcs))

            if ('FCS_WARNINGFUNCTION' in find_fcs )or ('AWV_WARNINGFUNCTION' in find_fcs):
                row_fcs = int(FCS_Position[len(FCS_Position) - 1])
                column_fcs_warning = self.FCS_Column + fcs
                print('Column FCS Warning function off = {0}'.format(column_fcs_warning))

            if 'BRAKEFUNCTION' in find_fcs:
                row_fcs = int(FCS_Position[len(FCS_Position) - 1])
                column_fcs_brake = self.FCS_Column + fcs
                print('Column FCS brake function off = {0}'.format(column_fcs_brake))

            if ('RADARSYSTEM_OFF' in find_fcs) or ('AWVSYSTEM_OFF' in find_fcs):
                row_fcs = int(FCS_Position[len(FCS_Position) - 1])
                column_fcs_system_off = self.FCS_Column + fcs
                print('Column FCS system off function off = {0}'.format(column_fcs_system_off))
            if (Video in find_fcs) and Video != '':
                row_fcs = int(FCS_Position[len(FCS_Position) - 1])
                column_fcs_Video = self.FCS_Column + fcs
                print('Column FCS Video function = {0}'.format(column_fcs_Video))
            # MAX_ROW = TestCaseFM.sheet.max_row + 1
        Position_Fault = int(First_Fault_Position[1])
        for i in range(Position_Fault, self.Max_Row_System_Degradation):
            reaction_fcs = []
            for j in range(FCS_Len):
                reaction_fcs.append(TestCaseFM.sheet.cell(row=i, column=self.FCS_Column + j).value)
            # print(len(reaction_fcs))
            Flag = ''
            for check_IR_RE in reaction_fcs:
                if check_IR_RE == 'IR' or check_IR_RE == 'ir':
                    Flag = 'IR'
                elif Flag != 'IR' and (check_IR_RE == 'RE' or check_IR_RE == 're'):
                    Flag = 'RE'
                elif (Flag != 'IR' and Flag != 'RE'):
                    Flag = ''

            Fault_Name_FCS = TestCaseFM.sheet[First_Fault_Position[0] + '{0}'.format(i)].value.strip()
            try:
                if Video == '':
                    if Flag == 'RE' \
                            and (reaction_fcs[column_fcs - self.FCS_Column] == 'RE' or reaction_fcs[column_fcs - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                    elif Flag == 'IR':
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_IRR_ALL'
                    elif Flag == 'RE' and (reaction_fcs[column_fcs - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs - self.FCS_Column] != 're') \
                            and ((reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_warning - self.FCS_Column] == 're') \
                                 or (reaction_fcs[column_fcs_system_off - self.FCS_Column] == 'RE' or reaction_fcs[
                                column_fcs_system_off - self.FCS_Column] == 're')) \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_brake - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_ALL'
                    elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_warning - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_system_off - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_brake - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_BRAKE'
                    elif (reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_warning - self.FCS_Column] == 're') \
                            and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_system_off - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_brake - self.FCS_Column] != 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_WARNING'
                    else:
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                elif Video != '':
                    if Flag == 'RE' \
                            and (reaction_fcs[column_fcs - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                    elif Flag == 'IR':
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_IRR_ALL'
                    elif Flag == 'RE' and (reaction_fcs[column_fcs - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs - self.FCS_Column] != 're') \
                            and ((reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_warning - self.FCS_Column] == 're') \
                                 or (reaction_fcs[column_fcs_system_off - self.FCS_Column] == 'RE' or reaction_fcs[
                                column_fcs_system_off - self.FCS_Column] == 're')) \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_brake - self.FCS_Column] == 're') \
                            and (reaction_fcs[column_fcs_Video - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_Video - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_ALL'
                    elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_warning - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_system_off - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_brake - self.FCS_Column] == 're') \
                            and (reaction_fcs[column_fcs_Video - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_Video - self.FCS_Column] != 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_BRAKE'
                    elif (reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_warning - self.FCS_Column] == 're') \
                            and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_system_off - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_brake - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_Video - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_Video - self.FCS_Column] != 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_WARNING'
                    elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_warning - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_system_off - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_brake - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_Video - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_Video - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_VIDEO'
                    elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_warning - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                        column_fcs_system_off - self.FCS_Column] != 're') \
                            and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_brake - self.FCS_Column] == 're') \
                            and (reaction_fcs[column_fcs_Video - self.FCS_Column] == 'RE' or reaction_fcs[
                        column_fcs_Video - self.FCS_Column] == 're'):
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_BRAKE_VIDEO'
                    else:
                        TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
            except:
                temp_fault = Fault_Name_FCS
                t_Red = re.search(r'(.*)FAULT_(.*)', temp_fault, re.M | re.I)
                if t_Red == None:
                    t_Red = re.search(r'(.*?)_(.*)', temp_fault, re.M | re.I)
                # print(t_Red.group(2))
                try:
                    for li_st in self.List_Fault:
                        if t_Red.group(2) in li_st[:]:
                            Fault_Name_FCS = li_st
                            break
                except:
                    print('Not find redundance {0}'.format(Fault_Name_FCS))
                # print(Fault_Name_FCS)
                try:
                    if Video == '':
                        if Flag == 'RE' \
                                and (reaction_fcs[column_fcs - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                        elif Flag == 'IR':
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_IRR_ALL'
                        elif Flag == 'RE' and (reaction_fcs[column_fcs - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs - self.FCS_Column] != 're') \
                                and ((reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_warning - self.FCS_Column] == 're') \
                                     or (reaction_fcs[column_fcs_system_off - self.FCS_Column] == 'RE' or reaction_fcs[
                                    column_fcs_system_off - self.FCS_Column] == 're')) \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_brake - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_ALL'
                        elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_warning - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_system_off - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_brake - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_BRAKE'
                        elif (reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_warning - self.FCS_Column] == 're') \
                                and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_system_off - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_brake - self.FCS_Column] != 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_WARNING'
                        else:
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                    elif Video != '':
                        if Flag == 'RE' \
                                and (reaction_fcs[column_fcs - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                        elif Flag == 'IR':
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_IRR_ALL'
                        elif Flag == 'RE' and (reaction_fcs[column_fcs - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs - self.FCS_Column] != 're') \
                                and ((reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_warning - self.FCS_Column] == 're') \
                                     or (reaction_fcs[column_fcs_system_off - self.FCS_Column] == 'RE' or reaction_fcs[
                                    column_fcs_system_off - self.FCS_Column] == 're')) \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_brake - self.FCS_Column] == 're') \
                                and (reaction_fcs[column_fcs_Video - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_Video - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_ALL'
                        elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_warning - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_system_off - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_brake - self.FCS_Column] == 're') \
                                and (reaction_fcs[column_fcs_Video - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_Video - self.FCS_Column] != 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_BRAKE'
                        elif (reaction_fcs[column_fcs_warning - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_warning - self.FCS_Column] == 're') \
                                and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_system_off - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_brake - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_Video - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_Video - self.FCS_Column] != 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_WARNING'
                        elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_warning - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_system_off - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_brake - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_Video - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_Video - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_VIDEO'
                        elif (reaction_fcs[column_fcs_warning - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_warning - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_system_off - self.FCS_Column] != 'RE' and reaction_fcs[
                            column_fcs_system_off - self.FCS_Column] != 're') \
                                and (reaction_fcs[column_fcs_brake - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_brake - self.FCS_Column] == 're') \
                                and (reaction_fcs[column_fcs_Video - self.FCS_Column] == 'RE' or reaction_fcs[
                            column_fcs_Video - self.FCS_Column] == 're'):
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_REV_BRAKE_VIDEO'
                        else:
                            TestCaseFM.DTCDict[str(Fault_Name_FCS)]['FUNCFCS'] = 'PSS_NoChange'
                except:
                    # print('Can not add {0} to TestCase'.format(Fault_Name_FCS))
                    pass

    def Time_Debounce(self, List_Fault=[], key_find=None):
        book1 = openpyxl.load_workbook(self)
        sheet1 = book1.active
        MAX_ROW_Book1 = sheet1.max_row + 1
        for fault in List_Fault:
            list_fautl = fault.split('_')
            re_fault = re.search(r'(.*?)_(.*?)_(.*)', fault, re.M | re.I)
            t1 = re_fault.group(1)
            t2 = re_fault.group(2)
            t3 = re_fault.group(3)
            Flag_max_min = False

            for i in range(2, MAX_ROW_Book1):
                Flag_confirm_type = 0  # 1 faulf is int, 2 find fault is int, 3 both fault are int
                if Flag_max_min == True:
                    break
                if i == MAX_ROW_Book1 - 1:
                    continue
                if sheet1.cell(row=i, column=3).value == 'Heading':
                    find_fault = sheet1.cell(row=i, column=2).value
                    list_find_fault = find_fault.split('_')
                    ct1 = 0
                    ct_position_fault = 0
                    ct_position_find_fault = 0
                    for t_key_find in key_find:
                        if t_key_find in list_find_fault:
                            for t_list_fautl in list_fautl:
                                ct_position_fault += 1
                                ct_position_find_fault = 0
                                for t_list_find_fault in list_find_fault:
                                    ct_position_find_fault += 1
                                    if (t_list_fautl.lower() in t_list_find_fault.lower()) and (ct_position_fault == ct_position_find_fault):
                                        ct1 += 1
                        else:
                            continue

                        try:
                            if isinstance(int(t_list_fautl[len(t_list_fautl) - 1]), int):
                                Flag_confirm_type = 1
                        except:
                            pass
                        try:
                            if isinstance(int(t_list_find_fault[len(t_list_find_fault) - 1]), int):
                                Flag_confirm_type = 2
                        except:
                            pass
                        try:
                            if isinstance(int(t_list_fautl[len(t_list_fautl) - 1]), int) and isinstance(
                                    int(t_list_find_fault[len(t_list_find_fault) - 1]), int):
                                Flag_confirm_type = 3
                        except:
                            pass

                        if ((t_key_find in list_fautl) and (ct1 == len(list_fautl))) or (
                                (t_key_find not in list_fautl) and (ct1 == (len(list_fautl) - 1))):
                            if Flag_confirm_type in [0, 3]:
                                j = i + 1
                                try:
                                    while 'Heading' not in sheet1.cell(row=j, column=3).value:
                                        # t1 = sheet1.cell(row=j, column=4)
                                        if ('Implemented' or 'Carry over') in sheet1.cell(row=j, column=4).value:
                                            min_max_time = sheet1.cell(row=j, column=2).value
                                            max_time = re.search(r'maxi\w*\s*\w*\s*(\:|.*?)(\d+)', min_max_time,
                                                                 re.M | re.I)
                                            min_time = re.search(r'mini\w*\s*\w*\s*(\:|.*?)(\d+)', min_max_time,
                                                                 re.M | re.I)
                                            qualification_time = re.search(r'qua\w*\s*\w*\s*(\:|.*?)(\d+)', min_max_time,
                                                                           re.M | re.I)
                                            if max_time is not None or min_time is not None:
                                                if max_time.group(2) == min_time.group(2):
                                                    TestCaseFM.DTCDict[str(fault)]['MAXTIME'] = max_time.group(2)
                                                    TestCaseFM.DTCDict[str(fault)]['MINTIME'] = str(int(min_time.group(2)) - 10)
                                                    Flag_max_min = True
                                                else:
                                                    TestCaseFM.DTCDict[str(fault)]['MAXTIME'] = max_time.group(2)
                                                    TestCaseFM.DTCDict[str(fault)]['MINTIME'] = min_time.group(2)
                                                    Flag_max_min = True
                                            elif qualification_time is not None:
                                                TestCaseFM.DTCDict[str(fault)]['MAXTIME'] = qualification_time.group(2)
                                                TestCaseFM.DTCDict[str(fault)]['MINTIME'] = str(
                                                    int(qualification_time.group(2)) - 10)
                                                Flag_max_min = True
                                        j += 1
                                    break
                                except:
                                    print('Fault not in Door: {0}'.format(fault))
                        # t_re_fault = re_fault.group(1) + '_' + t_key_find + '_' + re_fault.group(3)
                        # if t_re_fault in find_fault:

    def CFormWrite(self, testcase='', ExtraPF=False, ExtraSignal=False):

        if TestCaseFM.DTCDict[self]['FUNCFCO'] != '':
            if os.path.isfile(TestCaseFM.PathOfActivateFile + '\\{0}.py'.format(testcase.replace("FailureReports", "FailureReports_"))):
                if TestCaseFM.Tracking == 0:
                    TestCaseFM.Tracking = 1
                    Templatedtc = open(TestCaseFM.PathOfActivateFile + '\\{0}.py'.format(testcase.replace("FailureReports", "FailureReports_")), 'w')
                    Templatedtc.write('\n' * 2 + TestCaseFM.FromTem + '\n')
                    if ExtraPF:
                        Templatedtc.write(testcase + " = CFaultContainer('" + testcase + "')\n\n")
                        # Templatedtc.write(testcase[:14] + testcase[15:31] + testcase[32:34] + testcase[35:37] + " = CFaultContainer('" + \
                        #                   testcase[:14] + testcase[15:31] + testcase[32:34] + testcase[35:37] + "')\n\n")
                    else:
                        Templatedtc.write(
                            testcase + " = CFaultContainer('" + testcase + "')\n\n")
                    Templatedtc.close()
            else:
                TestCaseFM.Tracking = 1
                Templatedtc = open(TestCaseFM.PathOfActivateFile + '\\{0}.py'.format(testcase.replace("FailureReports", "FailureReports_")), 'w')
                Templatedtc.write('\n' * 2 + TestCaseFM.FromTem)
                if ExtraPF:
                    Templatedtc.write(testcase + " = CFaultContainer('" + testcase + "')\n\n")
                else:
                    Templatedtc.write(testcase + " = CFaultContainer('" + testcase + "')\n\n")
                Templatedtc.close()

            Templatedtc = open(TestCaseFM.PathOfActivateFile + '\\{0}.py'.format(testcase.replace("FailureReports", "FailureReports_")), 'a')
            Templatedtc.write('#----------------------------------------------------------------------\n')
            if ExtraPF:
                Templatedtc.write(testcase + '.append(name = ' + "'" + self + "'" + ',' + '\n')
                # Templatedtc.write(testcase[:14] + testcase[15:31] + testcase[32:34] + testcase[35:37] + \
                #                   '.append(name = ' + "'" + self + "'" + ',' + '\n')
                Templatedtc.write('\t' * 8 + 'comment=' + "''" + ',' + '\n')
                Templatedtc.write('\t' * 8 + "typ='" + 'PF_MM_SignalMonitoring' + "'" + ',' + '\n')
            else:
                Templatedtc.write(testcase + '.append(name = ' + "'" + self + "'" + ',' + '\n')
                Templatedtc.write('\t' * 8 + 'comment=' + "''" + ',' + '\n')
                if ExtraSignal:
                    Templatedtc.write('\t' * 8 + "typ='" + 'Compo_SignalMonitoring' + "'" + ',' + '\n')
                else:
                    Templatedtc.write('\t' * 8 + "typ='" + testcase[14:] + "'" + ',' + '\n')
            Templatedtc.write('\t' * 8 + "ReferenceDict={}," + '\n')
            if ExtraSignal:
                Templatedtc.write('\t' * 8 + "Conditions=['signal updated']," + '\n')
            else:
                Templatedtc.write('\t' * 8 + "Conditions=['']," + '\n')
            Templatedtc.write('\t' * 8 + "ExtendedOptions=['']," + '\n')
            Templatedtc.write('\t' * 8 + "CodingData={}," + '\n')
            Templatedtc.write('\t' * 8 + "preConditionsSeq=[]," + '\n')
            if ExtraPF:
                Templatedtc.write('\t' * 8 + "FailureCondition = [('XcpCalibrate.Dal_Test_ID_UW.setRawValue'," + \
                                  TestCaseFM.DTCDict[self]['Fault_ID'] + ")]," + '\n')
                Templatedtc.write('\t' * 8 + "Frame=''," + '\n')
                Templatedtc.write('\t' * 8 + "Pdu = ''," + '\n')
                Templatedtc.write('\t' * 8 + "Signal=''," + '\n')
            elif ExtraSignal:
                Templatedtc.write('\t' * 8 + "FailureCondition = [('TX_EMS_1.ECFail', 0x1)],\n")
                Templatedtc.write('\t' * 8 + "Frame=''," + '\n')
                Templatedtc.write('\t' * 8 + "Pdu = 'NotFound'," + '\n')
                Templatedtc.write('\t' * 8 + "Signal='NotFound'," + '\n')
            else:
                Templatedtc.write('\t' * 8 + "Frame=''," + '\n')
                if TestCaseFM.DTCDict[self]['MESSAGE'] == '':
                    Templatedtc.write('\t' * 8 + "Message='NotFound'," + '\n')
                else:
                    Templatedtc.write('\t' * 8 + "Message='" + TestCaseFM.DTCDict[self]['MESSAGE'] + "'," + '\n')

                if TestCaseFM.DTCDict[self]['SIGNAL'] == '':
                    if testcase != 'FailureReportsTimeout' and testcase != 'FailureReportsDLC':
                        Templatedtc.write('\t' * 8 + "Signal='NotFound'," + '\n')
                    else:
                        Templatedtc.write('\t' * 8 + "Signal=''," + '\n')
                else:
                    Templatedtc.write('\t' * 8 + "Signal='" + TestCaseFM.DTCDict[self]['SIGNAL'] + "'," + '\n')

            Templatedtc.write('\t' * 8 + "E2E={}," + '\n')
            if TestCaseFM.DTCDict[self]['MAXTIME'] != '':
                Templatedtc.write('\t' * 8 + "maxErrorDetectionTime=" + TestCaseFM.DTCDict[self]['MAXTIME'] + ',\n')
                Templatedtc.write('\t' * 8 + "minErrorDetectionTime=" + TestCaseFM.DTCDict[self]['MINTIME'] + ',\n')
            else:
                Templatedtc.write('\t' * 8 + "maxErrorDetectionTime=60," + '\n')
                Templatedtc.write('\t' * 8 + "minErrorDetectionTime=40," + '\n')
                if 'FAULT_TO' in self or 'FAULT_TIMEOUT' in self or 'FAULT_ALV' in self or 'FAULT_ALIVE' in self or 'RLC' in self \
                        or 'FAULT_CHK' in self or 'FAULT_CHECKSUM' in self or 'DLC' in self:
                    TestCaseFM.Redundance.write(str(self) + '\n')
                    print('This Fault: {0} have no max and min time'.format(self))
            Templatedtc.write('\t' * 8 + "DEM_debounceCounter=3," + '\n')
            Templatedtc.write('\t' * 8 + "debounceStepsFailurePresent=1," + '\n')
            Templatedtc.write('\t' * 8 + "debounceStepsFailureAbsent=3," + '\n')
            Templatedtc.write(
                '\t' * 8 + "systemFitnessState=[systemFitnessStates." + TestCaseFM.DTCDict[self][
                    'FUNCFCO'] + ", systemFitnessStates." \
                + TestCaseFM.DTCDict[self]['FUNCFCS'] + "]," + '\n')
            Templatedtc.write('\t' * 8 + "DTC=DTC." + TestCaseFM.DTCDict[self]['DTC'] + ",\n")
            Templatedtc.write('\t' * 8 + ")\n")
            Templatedtc.close()
        else:
            pass

    def collect_messNsig(self, dbcfilepath):
        file_dbc = open(dbcfilepath, mode='rt', buffering=1, encoding='iso8859_15')
        all_lines = file_dbc.readlines()
        message_package = {}
        flag_message_num = 0
        signals = []
        messages = []
        for line in all_lines:
            line_list = line.split()
            if line_list != [] and len(line_list) > 4:
                if line_list[0] == 'BO_':
                    messages.append(line_list[2].strip(':'))
                    if flag_message_num >= 1:
                        message_package[messages[flag_message_num - 1]] = signals
                        signals = []
                    flag_message_num = flag_message_num + 1
                elif line_list[0] == 'SG_':
                    signals.append(line_list[1])
        return message_package

    def find_message(self, t_message_package, t_fault_list):  # timeout, dlc, checksum, alive
        message_fault_map = {}
        t_fault_list = sorted(t_fault_list, key=len)
        for fault in t_fault_list:
            for message in t_message_package:
                if message.lower() in fault.lower():
                    message_fault_map[fault] = message
                    TestCaseFM.DTCDict[str(fault)]['MESSAGE'] = message
                    # print(TestCaseFM.DTCDict[str(fault)]['MESSAGE'])
                    break
                if 'vid' in fault.lower() and 'mounting' in fault.lower():
                    message_fault_map[fault] = 'VID_Mounting'
                    TestCaseFM.DTCDict[str(fault)]['MESSAGE'] = 'VID_Mounting'
                    # print(TestCaseFM.DTCDict[str(fault)]['MESSAGE'])
                    break
        return message_fault_map

    def find_signal(self, t_message_package, t_message_fault_map, t_fault_list, t_type):  # checksum, alive counter
        if t_type == 'checksum':
            name_samples = ['CheckSum', 'Chksm', 'Cksm']
        elif t_type == 'counter':
            name_samples = ['RollingCounter', 'Rolling_counter', 'MsgCnt', 'AlvRlngCtr', 'AlvRC', 'AliveRollCnt',
                            'AliveRC', 'ALIVE', 'Counter']
        else:
            name_samples = []

        signal_fault_map = {}

        signal_found = False
        for fault in t_fault_list:
            for sample in name_samples:
                message_name = t_message_fault_map.get(fault)
                if message_name:
                    if message_name == 'VID_Mounting':
                        if t_type == 'checksum':
                            signal_fault_map[fault] = 'VID_MountPos_ChkSum'
                            TestCaseFM.DTCDict[str(fault)]['SIGNAL'] = 'VID_MountPos_ChkSum'
                        elif t_type == 'counter':
                            signal_fault_map[fault] = 'VID_MountPos_MsgCnt'
                            TestCaseFM.DTCDict[str(fault)]['SIGNAL'] = 'VID_MountPos_MsgCnt'
                        else:
                            signal_fault_map[fault] = 'NotFound'
                        signal_found = True
                    else:
                        for signal in t_message_package[message_name]:
                            if sample.lower() in signal.lower():
                                signal_fault_map[fault] = signal
                                TestCaseFM.DTCDict[str(fault)]['SIGNAL'] = signal
                                # print(TestCaseFM.DTCDict[str(fault)]['SIGNAL'])
                                signal_found = True
                else:
                    break
                if signal_found:
                    signal_found = False
                    break

        return signal_fault_map

    def Get_ITC(self, path):
        myFile = open(path, "r")
        FileList = myFile.readlines()
        myFile.close

        FaultID = []
        for Str in FileList:
            tStr = Str
            tRegex = '(Fault\w+?)\W+?(\d+)u?$'
            tRegex2 = 'DemEventParameter_(\w+?)\W+?(\d+)u?$'
            if re.search(tRegex, tStr) is not None:
                result = re.search(tRegex, tStr).group()
                ID = int(re.sub(tRegex, r'\2', result))
                Name = re.sub(tRegex, r'\1', result)
                FaultID.append((ID, Name))
            elif re.search(tRegex2, tStr) is not None:
                result = re.search(tRegex2, tStr).group()
                ID = int(re.sub(tRegex2, r'\2', result))
                Name = re.sub(tRegex2, r'\1', result)
                FaultID.append((ID, Name))
        return FaultID

    def Modify_Fault_ID(self, Fault_ID):

        PFFile = open(TestCaseFM.PathOfActivateFile + '\\FailureReports_SignalMonitoring_PF.py', "r")
        PFData = PFFile.read()
        PFFile.close

        tRegex = '(name = \'(\w+)\'[^`]+?FailureCondition.*?=.*?\[\((.+?)\)\],)'
        list = re.findall(tRegex, PFData)
        ID = ""
        for tfault in list:
            for tFaultID in Fault_ID:
                if tFaultID[1] == tfault[1]:
                    ID = tFaultID[0]
                    break
            tStr = tfault[0].replace(tfault[2], "\'XcpCalibrate.Dal_Test_ID_UW.setRawValue\', " + str(ID))
            PFData = PFData.replace(tfault[0], tStr)

        PFFile = open(TestCaseFM.PathOfActivateFile + '\\FailureReports_SignalMonitoring_PF.py', "w")
        PFFile.write(PFData)
        PFFile.close

    def Main(self, Path_csv, PathSystemDegradation, PathTemplate, PathBook1, FCO_Position, First_Fault_Position, PathDBC, sheet_name, PathDTC, PathventID, Video='FCS_VRUWARNINGFUNCTION_OFF'):
        print('FCO position')
        print(FCO_Position)
        DTC.generate_DTC_py(PathDTC)
        t_FCO_len, t_FCS_len, t_FCS_Position = self.Openxlsx(PathSystemDegradation, First_Fault_Position, sheet_name, FCO_Position)
        self.Countcsv(Path_csv)
        self.GenDTCDict(Path_csv)
        self.ReadFormTem(PathTemplate)
        self.CalPosition(FCO_Position, t_FCS_Position)
        self.FCO(self, FCO_Position, t_FCO_len, First_Fault_Position)
        self.FCS(self, t_FCS_Position, t_FCS_len, First_Fault_Position, Video)

        for fault in self.List_Fault:
            if ('TIMEOUT' in fault[:] or 'TO_' in fault[:]) and 'CAN' not in fault[:]:
                self.List_Fault_Timeout.append(fault)
            elif 'ALIVE' in fault[:] or 'ALV_' in fault[:] or 'COUNTER' in fault[:] or 'RLC' in fault[:]:
                self.List_Fault_Alive.append(fault)
            elif 'CHECKSUM' in fault[:] or 'CHK_' in fault[:]:
                self.List_Fault_CheckSum.append(fault)
            elif 'DLC' in fault[:]:
                self.List_Fault_DLC.append(fault)
            else:
                for listPF in self.DataPF:
                    if listPF in fault[:]:
                        self.List_Fault_PF_MM.append(fault)
                        break
                else:
                    self.List_Fault_Signal.append(fault)

        tmp = self.collect_messNsig(self, PathDBC)

        self.find_message(self, tmp, self.List_Fault_Timeout)
        if self.List_Fault_DLC is not None:
            self.find_message(self, tmp, self.List_Fault_DLC)
        CHK_mess_m = self.find_message(self, tmp, self.List_Fault_CheckSum)
        ALV_mess_m = self.find_message(self, tmp, self.List_Fault_Alive)

        CHK_signal = self.find_signal(self, tmp, CHK_mess_m, self.List_Fault_CheckSum, 'checksum')
        ALV_signal = self.find_signal(self, tmp, ALV_mess_m, self.List_Fault_Alive, 'counter')

        print(len(self.List_Fault_Timeout), self.List_Fault_Timeout)
        print(len(self.List_Fault_Alive), self.List_Fault_Alive)
        print(len(self.List_Fault_CheckSum), self.List_Fault_CheckSum)
        print(len(self.List_Fault_DLC), self.List_Fault_DLC)
        print(len(self.List_Fault_PF_MM), self.List_Fault_PF_MM)
        print(len(self.List_Fault_Signal), self.List_Fault_Signal)

        self.Time_Debounce(PathBook1, self.List_Fault_Timeout, ['TO', 'TIMEOUT'])
        self.Time_Debounce(PathBook1, self.List_Fault_Alive, ['RLC', 'ALV', 'ALIVECOUNTER', 'ALIVE'])
        self.Time_Debounce(PathBook1, self.List_Fault_CheckSum, ['CHK', 'CHECKSUM'])
        self.Time_Debounce(PathBook1, self.List_Fault_DLC, ['DLC'])

        print(len(self.DTCDict), self.DTCDict)
        self.Redundance = open(TestCaseFM.PathOfActivateFile + '\\Redundance.py', 'a')
        TestCaseFM.Redundance.write(
            '\n' * 2 + '"===================Fault have no Maximum and Minimum Time====================="' + '\n')
        TestCaseFM.Redundance.write('"===================Please, Check Door again====================="' + '\n')
        if self.List_Fault_Timeout != '':
            self.Tracking = 0
            for fault in self.List_Fault_Timeout:
                self.CFormWrite(fault, 'FailureReportsTimeout', ExtraPF=False, ExtraSignal=False)
        if self.List_Fault_Alive != '':
            self.Tracking = 0
            for fault in self.List_Fault_Alive:
                self.CFormWrite(fault, 'FailureReportsAliveCounter', ExtraPF=False, ExtraSignal=False)
        if self.List_Fault_CheckSum != '':
            self.Tracking = 0
            for fault in self.List_Fault_CheckSum:
                self.CFormWrite(fault, 'FailureReportsChecksum', ExtraPF=False, ExtraSignal=False)
        if self.List_Fault_DLC != '':
            self.Tracking = 0
            for fault in self.List_Fault_DLC:
                self.CFormWrite(fault, 'FailureReportsDLC', ExtraPF=False, ExtraSignal=False)
        if self.List_Fault_PF_MM != '':
            self.Tracking = 0
            for fault in self.List_Fault_PF_MM:
                self.CFormWrite(fault, 'FailureReportsSignalMonitoring_PF', ExtraPF=True, ExtraSignal=False)
        if self.List_Fault_Signal != '':
            self.Tracking = 0
            for fault in self.List_Fault_Signal:
                self.CFormWrite(fault, 'FailureReportsSignalMonitoring', ExtraPF=False, ExtraSignal=True)
        if TestCaseFM.Confirm_PF_28 == True:
            List_fault_ITC = self.Get_ITC(self, PathventID)
            print(List_fault_ITC)
            self.Modify_Fault_ID(self, List_fault_ITC)
        TestCaseFM.Confirm_PF_28 = False
        self.Redundance.close()
        self.count = 0
        self.count2 = 0

# TestCaseFM.Main(TestCaseFM, TestCaseFM.Path_csv, TestCaseFM.PathSystemDegradation, TestCaseFM.PathTemplate,
#                 TestCaseFM.PathBook1, FCO_Position, TestCaseFM.FCO_Len, TestCaseFM.FCS_Position,
#                 TestCaseFM.FCS_Len, TestCaseFM.First_Fault_Position, TestCaseFM.PathDBC, TestCaseFM.sheet_name)
